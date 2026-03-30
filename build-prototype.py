import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Preschool Eval'

ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25
ws.page_margins.top = 0.25
ws.page_margins.bottom = 0.25

RED = 'CC1724'
DARK = '333333'
WHITE = 'FFFFFF'
LGRAY = 'F0F0F0'
MGRAY = 'E2E2E2'

thin = Side(style='thin', color='D0D0D0')
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal='center', vertical='center')
lft = Alignment(horizontal='left', vertical='center')
rot = Alignment(textRotation=45, horizontal='center', vertical='bottom')

red_fill = PatternFill('solid', fgColor=RED)
dark_fill = PatternFill('solid', fgColor=DARK)
lgray_fill = PatternFill('solid', fgColor=LGRAY)
white_fill = PatternFill('solid', fgColor=WHITE)
mgray_fill = PatternFill('solid', fgColor=MGRAY)

students = ['Elliana J.', 'Raiylah J.', 'Ellie T.', 'Adelyne Z.', 'Sofia M.', 'Harper L.']

skills = [
    ('VAULT', 'Run + Hurdle to Two-Foot Jump', ['Runs into hurdle, no stopping', 'Hurdles off one foot to both', 'Arms by ears in jump']),
    ('BARS', 'Front Support Fwd Roll (spot)', ['Unassisted front support 3 sec', 'Looks at belly while rolling', 'Hands stay on bar until complete']),
    ('BARS', 'Tuck Hang', ['Holds for 3 sec', 'Knees above belly button', 'Arms by ears']),
    ('BEAM', 'Walk Across Low Beam', ['Chest up, arms out to side', 'Steps heel to toe', 'Walks across without wobbles']),
    ('BEAM', 'Bear Crawl on Line', ['Hands on the line', 'Feet on the line', 'Straight legs']),
    ('BEAM', 'Backward Walk with Coach', ['Chest up, arms out to side', 'Steps heel to toe backward', 'Eyes forward while walking']),
    ('FLOOR', 'Monkey Jump Over Panel Mat', ['Turns hands to favorite foot', 'Jumps over without touching', 'Start & finish arms up']),
    ('FLOOR', 'Backward HS Walk Up (5 sec)', ['Walks feet up, hips over shoulders', 'Straight arms covering ears', 'Straight back']),
    ('FLOOR', 'Forward Roll Down Wedge', ['Starts in pencil shape', 'Eyes on belly', 'Stands up without hands']),
    ('FLOOR', 'Tabletop Hold', ['Holds for 5 sec', 'Back flat, no sagging hips', 'Fingers facing feet']),
    ('SAFETY', 'Safety', ['Follows directions', 'Stays with the group', 'Keeps hands to self']),
]

# Column widths: LEFT block (A-H), gap (I), RIGHT block (J-Q)
for col, w in {'A':4,'B':30,'C':7.5,'D':7.5,'E':7.5,'F':7.5,'G':7.5,'H':7.5,'I':1.5,'J':4,'K':30,'L':7.5,'M':7.5,'N':7.5,'O':7.5,'P':7.5,'Q':7.5}.items():
    ws.column_dimensions[col].width = w

# Row 1: Gym header
ws.merge_cells('A1:H1')
ws.merge_cells('J1:Q1')
for cell_ref in ['A1', 'J1']:
    c = ws[cell_ref]
    c.value = 'HGA  ·  Houston Gymnastics Academy  ·  Preschool  ·  Mon 3:30'
    c.font = Font(name='Arial', bold=True, size=9, color=WHITE)
    c.fill = red_fill
    c.alignment = ctr
ws.row_dimensions[1].height = 22

# Row 2: Student names
ws.merge_cells('A2:B2')
ws.merge_cells('J2:K2')
ws['A2'].value = 'Skill / Criteria'
ws['A2'].font = Font(name='Arial', size=7, color='888888')
ws['A2'].alignment = lft
ws['J2'].value = 'Skill / Criteria'
ws['J2'].font = Font(name='Arial', size=7, color='888888')
ws['J2'].alignment = lft

for i, name in enumerate(students):
    for base_col in [3, 12]:  # C and L
        c = ws.cell(row=2, column=base_col + i)
        c.value = name
        c.font = Font(name='Arial', bold=True, size=7)
        c.alignment = rot
        c.fill = mgray_fill
ws.row_dimensions[2].height = 50


def write_event_band(row, start_col, end_col, event_name):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col)
    c.value = event_name
    c.font = Font(name='Arial', bold=True, size=8, color=WHITE)
    c.fill = red_fill
    c.alignment = ctr


def write_skill_block(row, start_col, skill_name, criteria):
    end_col = start_col + 7
    # Title bar
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col)
    c.value = skill_name
    c.font = Font(name='Arial', bold=True, size=8, color=WHITE)
    c.fill = dark_fill
    c.alignment = lft
    ws.row_dimensions[row].height = 17

    # Criteria rows
    for ci, crit in enumerate(criteria):
        r = row + 1 + ci
        bg = lgray_fill if ci % 2 == 1 else white_fill

        c = ws.cell(row=r, column=start_col)
        c.value = f'\u2605{ci+1}'
        c.font = Font(name='Arial', bold=True, size=9, color=RED)
        c.alignment = ctr
        c.fill = bg
        c.border = bdr

        c = ws.cell(row=r, column=start_col + 1)
        c.value = crit
        c.font = Font(name='Arial', size=8)
        c.alignment = lft
        c.fill = bg
        c.border = bdr

        for si in range(6):
            c = ws.cell(row=r, column=start_col + 2 + si)
            c.value = '\u2606'
            c.font = Font(name='Arial', size=16, color='CCCCCC')
            c.alignment = ctr
            c.fill = bg
            c.border = bdr

        ws.row_dimensions[r].height = 22

    # Mastery row
    r = row + len(criteria) + 1
    c = ws.cell(row=r, column=start_col)
    c.value = '\u2605'
    c.font = Font(name='Arial', bold=True, size=10, color='DAA520')
    c.alignment = ctr
    c.fill = white_fill
    c.border = bdr

    c = ws.cell(row=r, column=start_col + 1)
    c.value = '3x in a row'
    c.font = Font(name='Arial', size=7, color='999999', italic=True)
    c.alignment = lft
    c.fill = white_fill
    c.border = bdr

    for si in range(6):
        c = ws.cell(row=r, column=start_col + 2 + si)
        c.value = '\u2606'
        c.font = Font(name='Arial', size=16, color='DDDDDD')
        c.alignment = ctr
        c.fill = white_fill
        c.border = bdr

    ws.row_dimensions[r].height = 22


# Layout skills two-up
row = 3
skill_idx = 0
last_event = ['', '']

while skill_idx < len(skills):
    left = skills[skill_idx]
    right = skills[skill_idx + 1] if skill_idx + 1 < len(skills) else None

    # Event bands
    if left[0] != last_event[0]:
        write_event_band(row, 1, 8, left[0])
        last_event[0] = left[0]
    if right and right[0] != last_event[1]:
        write_event_band(row, 10, 17, right[0])
        last_event[1] = right[0]
    elif right and right[0] == last_event[1]:
        # Same event, still need the band for visual consistency
        write_event_band(row, 10, 17, right[0])

    ws.row_dimensions[row].height = 16
    row += 1

    # Skill blocks
    write_skill_block(row, 1, left[1], left[2])
    if right:
        write_skill_block(row, 10, right[1], right[2])

    row += 5  # title + 3 criteria + mastery
    skill_idx += 2

# Footer
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
c = ws.cell(row=row, column=1)
c.value = '\u2605 red = criterion passed  \u00b7  \u2605 gold = all criteria mastered 3\u00d7 in a row'
c.font = Font(name='Arial', size=6, color='999999')

ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=17)
c = ws.cell(row=row, column=10)
c.value = 'Skills & Stars \u00b7 Preschool \u00b7 Houston Gymnastics Academy'
c.font = Font(name='Arial', size=6, color='999999')
c.alignment = Alignment(horizontal='right')

output = 'C:/Users/Jayme/Downloads/eval-grid-twoup-clean.xlsx'
wb.save(output)
print(f'Saved to {output}')
